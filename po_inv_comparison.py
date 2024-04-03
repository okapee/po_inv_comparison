import sys
import json
import base64
import xml.etree.ElementTree as ET
import csv
from difflib import SequenceMatcher
import requests
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
    QFrame,
    QPushButton,
    QTextEdit,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QFileDialog,
    QGridLayout,
    QListWidget,
    QMessageBox,
    QProgressBar,
    QCheckBox,
    QTreeWidget,
    QTreeWidgetItem,
    QDialog,
    QLineEdit,
    QDialogButtonBox,
)
from PyQt5.QtGui import QFont, QColor, QPalette, QIcon
from PyQt5.QtCore import Qt, QThread, pyqtSignal
import openpyxl


class ComparisonApp(QMainWindow):
    _instance = None

    def __init__(self):
        super().__init__()
        print("ComparisonApp.__init__ called.")  # デバッグ用の出力
        self.initUI()
        self._csv_file = None
        self._xml_file = None
        self._xml_data = None

    @classmethod
    def instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    @property
    def csv_file(self):
        return self._csv_file

    @csv_file.setter
    def csv_file(self, value):
        self._csv_file = value

    @property
    def xml_data(self):
        return self._xml_data

    @xml_data.setter
    def xml_data(self, value):
        self._xml_data = value

    def initUI(self):
        print("ComparisonApp.initUI called.")  # デバッグ用の出力
        self.setWindowTitle("発注書とPeppolの照合")
        self.setGeometry(100, 100, 1200, 800)  # ウィンドウサイズを変更
        self.setStyleSheet(
            "background-color: #FFFFFF; color: #333333; font-family: 'Arial';"  # フォントファミリを指定
        )

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        title_label = QLabel("発注書とPeppolの照合")
        title_label.setFont(QFont("Arial", 24, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #1A237E; margin-bottom: 20px;")
        layout.addWidget(title_label)

        drop_layout = QGridLayout()
        layout.addLayout(drop_layout)

        self.csv_drop_label = self.create_drop_label("📜", "発注書をここにドロップ")
        self.xml_drop_label = self.create_drop_label("💻", "Peppolをここにドロップ")

        drop_layout.addWidget(self.csv_drop_label, 0, 0)
        drop_layout.addWidget(self.xml_drop_label, 0, 1)

        self.call_api_button = QPushButton("Peppol APからインボイスを取得")
        self.call_api_button.clicked.connect(self.call_api)
        self.call_api_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        drop_layout.addWidget(self.call_api_button, 1, 1)

        self.clear_button = QPushButton("発注書とPeppolをリセット")
        self.clear_button.clicked.connect(self.clear_data)
        self.clear_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        layout.addWidget(self.clear_button, alignment=Qt.AlignRight)

        self.execute_button = QPushButton("照合実行")
        self.execute_button.clicked.connect(self.compare_files)
        self.execute_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        layout.addWidget(self.execute_button, alignment=Qt.AlignCenter)

        # self.result_table = QTableWidget()
        # self.result_table.setColumnCount(3)  # 列数を3に変更
        # self.result_table.setHorizontalHeaderLabels(["ステータス", "発注書", "Peppol"])
        # self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # self.result_table.verticalHeader().setVisible(False)
        # self.result_table.setStyleSheet("font-size: 14px;")
        # layout.addWidget(self.result_table)
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(3)  # 列数を3に変更
        self.result_table.setHorizontalHeaderLabels(["ステータス", "発注書", "Peppol"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setStyleSheet("font-size: 14px;")
        self.result_table.setMinimumHeight(200)  # 最小の高さを設定
        layout.addWidget(self.result_table)

        self.export_button = QPushButton("Excelで出力")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        self.export_button.setEnabled(False)
        layout.addWidget(self.export_button, alignment=Qt.AlignCenter)

        # xml_fileをNoneに初期化
        self.xml_file = None

    def create_drop_label(self, icon, text):
        label = QWidget()
        label.setAcceptDrops(True)
        label.dragEnterEvent = self.dragEnterEvent
        label.dragMoveEvent = self.dragMoveEvent
        label.dropEvent = self.dropEvent
        label.setStyleSheet(
            "background-color: #F5F5F5; border: 2px #1A237E; border-radius: 5px; padding: 20px; font-size: 16px;"
        )

        icon_label = QLabel(icon)
        icon_label.setFont(QFont("Arial", 24))

        text_label = QLabel(text)
        text_label.setObjectName("text_label")

        layout = QHBoxLayout()
        layout.addWidget(icon_label)
        layout.addWidget(text_label)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)

        label.setLayout(layout)

        return label

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.endswith(".csv"):
                    self._csv_file = file_path  # csv_fileを直接設定
                    self.csv_drop_label.findChild(QLabel, "text_label").setText(
                        f"発注書: {file_path.split('/')[-1]}"
                    )
                elif file_path.endswith(".xml"):
                    self._xml_file = file_path  # xml_fileを直接設定
                    self.xml_drop_label.findChild(QLabel, "text_label").setText(
                        f"Peppol: {file_path.split('/')[-1]}"
                    )
                    self.call_api_button.setEnabled(False)

    def call_api(self):
        if not self.csv_file:
            QMessageBox.warning(self, "警告", "先に発注書を選択してください。")
            return
        self.invoice_window = InvoiceWindow(self)
        self.invoice_window.show()

    def clear_data(self):
        self.csv_file = None
        self.xml_file = None
        self.xml_data = None
        self.csv_drop_label.findChild(QLabel, "text_label").setText(
            "発注書をここにドロップ"
        )
        self.xml_drop_label.findChild(QLabel, "text_label").setText(
            "Peppolをここにドロップ"
        )
        self.xml_drop_label.setEnabled(True)
        self.call_api_button.setEnabled(True)
        self.result_table.setRowCount(0)
        self.export_button.setEnabled(False)

    def compare_files(self):
        if (self.csv_file and self.xml_file) or (self.csv_file and self.xml_data):
            try:
                if self.xml_file:
                    xml_data = self.parse_xml(self.xml_file)
                else:
                    xml_data = self.xml_data
                csv_data = self.parse_csv(self.csv_file)
                diff_result = self.compare_data(xml_data, csv_data)

                self.result_table.setRowCount(len(diff_result))
                for row, item in enumerate(diff_result):
                    status_item = QTableWidgetItem(item["status"])
                    if item["status"] == "発注書のみに存在":
                        status_item.setForeground(QColor("green"))
                    elif item["status"] == "Peppolのみに存在":
                        status_item.setForeground(QColor("red"))
                    elif item["status"] == "差分あり":
                        status_item.setForeground(QColor("blue"))
                    self.result_table.setItem(row, 0, status_item)
                    self.result_table.setItem(row, 1, QTableWidgetItem(item["key"]))
                    self.result_table.setItem(
                        row, 2, QTableWidgetItem(item["csv_value"])
                    )
                    self.result_table.setItem(
                        row, 3, QTableWidgetItem(item["xml_value"])
                    )

                self.export_button.setEnabled(True)

            except (ValueError, ET.ParseError) as e:
                print(f"エラー: {e}")
        else:
            print(
                "発注書とPeppolをドロップまたはPeppol APからインボイスを取得してください"
            )

    def compare_files_for_array(self):
        print("compare_files_for_array called.")
        if self.csv_file and self.xml_data:
            try:
                xml_data = self.xml_data
                csv_data = self.parse_csv(self.csv_file)
                print("XML Data:")  # デバッグ用の出力
                print(xml_data)  # デバッグ用の出力
                print("CSV Data:")  # デバッグ用の出力
                print(csv_data)  # デバッグ用の出力
                diff_result = self.compare_data_for_array(xml_data, csv_data)
                print("Diff Result:")  # デバッグ用の出力
                print(diff_result)  # デバッグ用の出力

                print(f"result_table:{self.result_table}")

                self.result_table.setRowCount(0)  # 既存の行をクリア
                self.result_table.setRowCount(len(diff_result))
                for row, item in enumerate(diff_result):
                    status_item = QTableWidgetItem(item["status"])
                    if item["status"] == "発注書のみに存在":
                        status_item.setForeground(QColor("green"))
                    elif item["status"] == "Peppolのみに存在":
                        status_item.setForeground(QColor("red"))
                    elif item["status"] == "差分あり":
                        status_item.setForeground(QColor("blue"))
                    self.result_table.setItem(row, 0, status_item)
                    self.result_table.setItem(
                        row, 1, QTableWidgetItem(item["csv_value"])
                    )
                    self.result_table.setItem(
                        row, 2, QTableWidgetItem(item["xml_value"])
                    )

                print("Result Table:")  # デバッグ用の出力
                for row in range(
                    self.result_table.rowCount()
                ):  # 各行についてデバッグ出力
                    row_data = []
                    for col in range(self.result_table.columnCount()):
                        item = self.result_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    print(f"Row {row}: {row_data}")

                self.export_button.setEnabled(True)

            except (ValueError, ET.ParseError) as e:
                print(f"エラー: {e}")
        else:
            print(
                "発注書とPeppolをドロップまたはPeppol APからインボイスを取得してください"
            )

    def compare_files_with_invoice_data(self, invoice_data):
        print(f"csv_file in invoice_data:{self.csv_file}")
        print(f"invoice_data in invoice_data:{invoice_data}")
        if self.csv_file and invoice_data:
            try:
                csv_data = self.parse_csv(self.csv_file)
                print("CSV Data:")  # デバッグ用の出力
                print(csv_data)  # デバッグ用の出力
                diff_result = self.compare_data_for_array(invoice_data, csv_data)
                print("Diff Result:")  # デバッグ用の出力
                print(diff_result)  # デバッグ用の出力

                # 新しいComparisonAppのウィンドウを作成
                result_window = ComparisonApp()
                result_window.result_table.setRowCount(0)  # 既存の行をクリア
                result_window.result_table.setColumnCount(3)  # 列数を3に設定
                result_window.result_table.setHorizontalHeaderLabels(
                    ["ステータス", "発注書", "Peppol"]
                )
                result_window.result_table.setRowCount(len(diff_result))
                for row, item in enumerate(diff_result):
                    status_item = QTableWidgetItem(item["status"])
                    if item["status"] == "発注書のみに存在":
                        status_item.setForeground(QColor("green"))
                    elif item["status"] == "Peppolのみに存在":
                        status_item.setForeground(QColor("red"))
                    result_window.result_table.setItem(row, 0, status_item)
                    result_window.result_table.setItem(
                        row, 1, QTableWidgetItem(item["csv_value"])
                    )
                    result_window.result_table.setItem(
                        row, 2, QTableWidgetItem(item["xml_value"])
                    )

                print("Result Table:")  # デバッグ用の出力
                for row in range(
                    result_window.result_table.rowCount()
                ):  # 各行についてデバッグ出力
                    row_data = []
                    for col in range(result_window.result_table.columnCount()):
                        item = result_window.result_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    print(f"Row {row}: {row_data}")

                result_window.export_button.setEnabled(True)
                result_window.show()  # 新しいウィンドウを表示
                self.close()  # 元のComparisonAppのウィンドウを閉じる

            except (ValueError, ET.ParseError) as e:
                print(f"エラー: {e}")
        else:
            print("発注書とインボイスデータを選択してください")

    def compare_data_for_array(self, xml_data, csv_data):
        diff_result = []
        csv_keys = [
            (row["Description"], row["Quantity"], row["Price"]) for row in csv_data
        ]
        xml_keys = [
            (row["item_name"], row["quantity"], row["price"]) for row in xml_data
        ]

        for csv_key in csv_keys:
            if not any(self.is_similar_tuple(csv_key, xml_key) for xml_key in xml_keys):
                diff_result.append(
                    {
                        "status": "発注書のみに存在",
                        "csv_value": ", ".join(
                            map(str, csv_key)
                        ),  # タプルを文字列に変換
                        "xml_value": "",
                    }
                )

        for xml_key in xml_keys:
            if not any(self.is_similar_tuple(xml_key, csv_key) for csv_key in csv_keys):
                diff_result.append(
                    {
                        "status": "Peppolのみに存在",
                        "csv_value": "",
                        "xml_value": ", ".join(
                            map(str, xml_key)
                        ),  # タプルを文字列に変換
                    }
                )

        return diff_result

    def parse_xml(self, xml_file):
        # XMLファイルをパースしてルート要素を取得
        tree = ET.parse(xml_file)
        root = tree.getroot()

        xml_data = []
        namespaces = {
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        }
        for line in root.findall(".//cac:InvoiceLine", namespaces=namespaces):
            item_name = line.find(".//cac:Item/cbc:Name", namespaces=namespaces).text
            quantity = line.find(".//cbc:InvoicedQuantity", namespaces=namespaces).text
            price = line.find(
                ".//cac:Price/cbc:PriceAmount", namespaces=namespaces
            ).text
            xml_data.append(
                {"item_name": item_name, "quantity": quantity, "price": price}
            )

        return xml_data

    def parse_csv(self, csv_file):
        csv_data = []
        with open(csv_file, "r", encoding="utf-8") as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                csv_data.append(row)

        return csv_data

    def compare_data(self, xml_data, csv_data):
        diff_result = []
        csv_keys = [tuple(row.values()) for row in csv_data]
        xml_keys = [
            (row["item_name"], row["quantity"], row["price"]) for row in xml_data
        ]

        for csv_key in csv_keys:
            if not any(self.is_similar_tuple(csv_key, xml_key) for xml_key in xml_keys):
                diff_result.append(
                    {
                        "status": "発注書のみに存在",
                        "key": "項目",
                        "csv_value": ", ".join(csv_key),
                        "xml_value": "",
                    }
                )

        for xml_key in xml_keys:
            if not any(self.is_similar_tuple(xml_key, csv_key) for csv_key in csv_keys):
                diff_result.append(
                    {
                        "status": "Peppolのみに存在",
                        "key": "項目",
                        "csv_value": "",
                        "xml_value": ", ".join(xml_key),
                    }
                )

        for csv_key, xml_key in zip(csv_keys, xml_keys):
            if self.is_similar_tuple(csv_key, xml_key):
                if csv_key != xml_key:
                    diff_result.append(
                        {
                            "status": "差分あり",
                            "key": "項目",
                            "csv_value": ", ".join(csv_key),
                            "xml_value": ", ".join(xml_key),
                        }
                    )

        return diff_result

    def is_similar_tuple(self, tuple1, tuple2, threshold=0.8):
        return all(
            self.is_similar(str1, str2, threshold) for str1, str2 in zip(tuple1, tuple2)
        )

    def is_similar(self, str1, str2, threshold=0.8):
        if str1.isdigit() and str2.isdigit():
            return str1 == str2
        else:
            return SequenceMatcher(None, str1, str2).ratio() >= threshold

    def export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Excelで保存", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "照合結果"

            # ヘッダーを書き込む
            sheet.cell(row=1, column=1, value="ステータス")
            sheet.cell(row=1, column=2, value="項目")
            sheet.cell(row=1, column=3, value="発注書")
            sheet.cell(row=1, column=4, value="Peppol")

            # 照合結果を書き込む
            for row in range(self.result_table.rowCount()):
                for col in range(self.result_table.columnCount()):
                    cell_value = self.result_table.item(row, col).text()
                    sheet.cell(row=row + 2, column=col + 1, value=cell_value)

            workbook.save(file_path)

    def select_csv_file(self):
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("CSV Files (*.csv)")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.csv_file = file_path


class InvoiceDetailsWindow(QMainWindow):
    def __init__(self, invoice_details_dict):
        super().__init__()
        self.setWindowTitle("選択した請求書の詳細")
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet(
            "background-color: #FFFFFF; color: #333333; font-family: 'Arial';"
        )

        # ComparisonAppのインスタンスを作成
        self.comparison_app = ComparisonApp.instance()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        title_label = QLabel("Invoice Details")
        title_label.setFont(QFont("Arial", 24, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #1A237E; margin-bottom: 20px;")
        layout.addWidget(title_label)

        self.tree_widget = QTreeWidget()
        self.tree_widget.setHeaderLabels(["", "Invoice ID", "品名", "数量", "単価"])
        self.tree_widget.setColumnWidth(0, 50)  # チェックボックスの列幅を調整
        self.tree_widget.setColumnWidth(1, 200)  # Invoice IDの列幅を設定
        layout.addWidget(self.tree_widget)

        for invoice_id, invoice_details in invoice_details_dict.items():
            invoice_item = QTreeWidgetItem([None, invoice_id])
            invoice_item.setFlags(invoice_item.flags() | Qt.ItemIsUserCheckable)
            invoice_item.setCheckState(0, Qt.Unchecked)
            for item in invoice_details["items"]:
                item_widget = QTreeWidgetItem(
                    [
                        None,
                        None,
                        item["item_name"],
                        str(item["quantity"]),
                        str(item["price"]),
                    ]
                )
                invoice_item.addChild(item_widget)
            self.tree_widget.addTopLevelItem(invoice_item)

        self.execute_button = QPushButton("選択したインボイスで消込実行")
        self.execute_button.clicked.connect(self.execute_selected_invoices)
        self.execute_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        layout.addWidget(self.execute_button, alignment=Qt.AlignCenter)

    def execute_selected_invoices(self):
        selected_invoice_items = []
        for i in range(self.tree_widget.topLevelItemCount()):
            invoice_item = self.tree_widget.topLevelItem(i)
            if invoice_item.checkState(0) == Qt.Checked:
                invoice_id = invoice_item.text(1)
                for j in range(invoice_item.childCount()):
                    item_widget = invoice_item.child(j)
                    item_name = item_widget.text(2)
                    quantity = item_widget.text(3)
                    price = item_widget.text(4)
                    selected_invoice_items.append(
                        {"item_name": item_name, "quantity": quantity, "price": price}
                    )

        print("Selected Invoice Items:")  # デバッグ用の出力
        print(selected_invoice_items)  # デバッグ用の出力

        if selected_invoice_items:
            self.comparison_app.activateWindow()  # ComparisonAppをアクティブにする

            # ComparisonAppのcsvファイルのパスを設定する
            csv_file_dialog = CSVFileDialog(self)

            if csv_file_dialog.exec_():
                csv_file_path = csv_file_dialog.get_file_path()
                self.comparison_app.csv_file = csv_file_path
                print("Calling compare_files_with_invoice_data...")  # デバッグ用の出力
                self.comparison_app.compare_files_with_invoice_data(
                    selected_invoice_items
                )
                print("compare_files_with_invoice_data called.")  # デバッグ用の出力
            else:
                QMessageBox.warning(self, "警告", "発注書がロードされていません。")

            self.close()  # InvoiceDetailsWindowを閉じる
        else:
            QMessageBox.warning(self, "警告", "選択されたインボイスがありません。")


# class CustomFileDialog(QFileDialog):
#     def __init__(self, *args, **kwargs):
#         super().__init__(*args, **kwargs)
#         self.setStyleSheet(
#             "background-color: #FFFFFF; color: #333333; font-family: 'Arial';"
#         )
#         self.setGeometry(100, 100, 800, 600)


class CSVFileDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("発注書のCSVを選択してください。")
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet(
            "background-color: #F5F5F5; color: #333333; font-family: 'Arial';"
        )

        layout = QVBoxLayout(self)

        title_label = QLabel("発注書のCSVファイルを選択してください。")
        title_label.setFont(QFont("Arial", 24, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #1A237E; margin-bottom: 20px;")
        layout.addWidget(title_label)

        file_path_label = QLabel("ファイルパス:")
        file_path_label.setFont(QFont("Arial", 14))
        file_path_label.setStyleSheet("margin-bottom: 5px;")
        layout.addWidget(file_path_label)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setFont(QFont("Arial", 14))
        self.file_path_edit.setStyleSheet(
            "background-color: #FFFFFF; border: 1px solid #1A237E; border-radius: 5px; padding: 5px;"
        )
        layout.addWidget(self.file_path_edit)

        browse_button = QPushButton("ファイルを選択")
        browse_button.setFont(QFont("Arial", 16))
        browse_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px; margin-top: 20px;"
        )
        browse_button.clicked.connect(self.browse_file)
        layout.addWidget(browse_button)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.setFont(QFont("Arial", 16))
        button_box.setStyleSheet(
            "background-color: #FFFFFF; color: #1A237E; font-size: 16px; padding: 10px 20px; border-radius: 5px; margin-top: 20px;"
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def browse_file(self):
        file_dialog = QFileDialog(self)
        file_dialog.setNameFilter("CSV Files (*.csv)")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_():
            file_path = file_dialog.selectedFiles()[0]
            self.file_path_edit.setText(file_path)

    def get_file_path(self):
        return self.file_path_edit.text()


class InvoiceWindow(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()
        self.invoice_details = {}  # invoice_details属性を初期化
        self.invoice_details_dict = {}

    def initUI(self):
        self.setWindowTitle("Peppolインボイス取得")
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet(
            "background-color: #FFFFFF; color: #333333; font-family: 'Arial';"  # フォントファミリを指定
        )

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        self.call_api_button = QPushButton("APIをコール")
        self.call_api_button.clicked.connect(self.call_api)
        self.call_api_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        layout.addWidget(self.call_api_button, alignment=Qt.AlignCenter)

        self.invoice_table = QTableWidget()
        self.invoice_table.setColumnCount(5)
        self.invoice_table.setHorizontalHeaderLabels(
            ["", "Invoice ID", "Sender ID", "Receiver ID", "Delivered Time"]
        )
        self.invoice_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.invoice_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.invoice_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )
        self.invoice_table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.Stretch
        )
        self.invoice_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.Stretch
        )
        self.invoice_table.verticalHeader().setVisible(False)
        self.invoice_table.setStyleSheet("font-size: 14px;")
        layout.addWidget(self.invoice_table)

        self.get_xml_button = QPushButton("選択したInvoiceのXMLを取得")
        self.get_xml_button.clicked.connect(self.get_selected_invoices_xml)
        self.get_xml_button.setStyleSheet(
            "background-color: #1A237E; color: #FFFFFF; font-size: 16px; padding: 10px 20px; border-radius: 5px;"
        )
        layout.addWidget(self.get_xml_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)  # インデeterminate モード
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

    def call_api(self):
        self.call_api_button.setEnabled(False)  # APIコールボタンを無効化
        self.progress_bar.setVisible(True)  # ローディングインジケーターを表示

        url = "https://api.prerelease.fa-peppol.com/api/v1.0/document/inbound/unread"
        token = "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxIiwiaWF0IjoxNjU3MDA5NjMzLCJleHAiOjE3MjgwMDAwMDB9.d610N8S3dpGNV-mht6uXkriC-iYIYQWF-BPQLlDKioPsWwqmvBCjLIUytsZG7HVGgwLGWEO6BInqdD6IzwA9aA"  # 実際のTokenに置き換えてください

        self.api_thread = APICallThread(url, token)
        self.api_thread.api_result.connect(self.handle_api_result)
        self.api_thread.start()

    def handle_api_result(self, invoices, error):
        self.progress_bar.setVisible(False)  # ローディングインジケーターを非表示
        self.call_api_button.setEnabled(True)  # APIコールボタンを有効化

        if error:
            QMessageBox.critical(self, "エラー", str(error))
        else:
            self.invoice_table.setRowCount(len(invoices))  # 行数を設定

            for row, invoice in enumerate(invoices):
                checkbox_item = QTableWidgetItem()
                checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                checkbox_item.setCheckState(Qt.Unchecked)
                self.invoice_table.setItem(row, 0, checkbox_item)

                invoice_id = invoice.get("id", "")
                peppol_sender_id = invoice.get("peppol_sender_id", "")
                peppol_receiver_id = invoice.get("peppol_receiver_id", "")
                delivered_time = invoice.get("delivered_time", "")

                self.invoice_table.setItem(row, 1, QTableWidgetItem(invoice_id))
                self.invoice_table.setItem(row, 2, QTableWidgetItem(peppol_sender_id))
                self.invoice_table.setItem(row, 3, QTableWidgetItem(peppol_receiver_id))
                self.invoice_table.setItem(row, 4, QTableWidgetItem(delivered_time))

    # 選択したInvoiceのXMLを取得 ボタンを押した際の処理
    def get_selected_invoices_xml(self):
        selected_invoice_ids = []
        for row in range(self.invoice_table.rowCount()):
            checkbox_item = self.invoice_table.item(row, 0)
            if checkbox_item.checkState() == Qt.Checked:
                invoice_id_item = self.invoice_table.item(row, 1)
                invoice_id = invoice_id_item.text()
                selected_invoice_ids.append(invoice_id)

        self.invoice_details_dict = {}  # invoice_details_dictを初期化
        self.pending_api_calls = len(
            selected_invoice_ids
        )  # 保留中のAPIコールの数を設定

        self.api_threads = []  # APIスレッドを格納するリストを追加

        for invoice_id in selected_invoice_ids:
            url = f"https://api.prerelease.fa-peppol.com/api/v1.0/document/inbound/{invoice_id}"
            token = "eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxIiwiaWF0IjoxNjU3MDA5NjMzLCJleHAiOjE3MjgwMDAwMDB9.d610N8S3dpGNV-mht6uXkriC-iYIYQWF-BPQLlDKioPsWwqmvBCjLIUytsZG7HVGgwLGWEO6BInqdD6IzwA9aA"  # 実際のTokenに置き換えてください

            api_thread = SelectedAPICallThread(url, token)
            api_thread.api_success.connect(
                lambda result, invoice_id=invoice_id: self.handle_api_success(
                    result, invoice_id
                )
            )
            api_thread.api_error.connect(self.handle_api_error)
            api_thread.start()

            self.api_threads.append(api_thread)  # APIスレッドをリストに追加

        self.wait_for_api_threads()  # APIスレッドの完了を待つ

    def wait_for_api_threads(self):
        for thread in self.api_threads:
            thread.wait()

        if self.pending_api_calls == 0:  # 全てのAPIコールが完了したら
            self.invoice_details_window = InvoiceDetailsWindow(
                self.invoice_details_dict
            )
            self.invoice_details_window.show()
            self.close()  # InvoiceWindowを閉じる

    def handle_api_success(self, result, invoice_id):
        invoice_details = self.handle_base64_xml_result(result, invoice_id)
        self.invoice_details_dict[invoice_id] = invoice_details
        self.pending_api_calls -= 1  # 保留中のAPIコールの数を減らす

        if self.pending_api_calls == 0:  # 全てのAPIコールが完了したら
            self.invoice_details_window = InvoiceDetailsWindow(
                self.invoice_details_dict
            )
            self.invoice_details_window.show()
            self.close()  # InvoiceWindowを閉じる

    def handle_xml_result(self, result, invoice_id):
        xml_data = result["xml_data"]
        print("XML Data:")  # デバッグ用の出力
        print(xml_data)  # デバッグ用の出力
        invoice_details = self.parse_xml(xml_data)
        self.invoice_details[invoice_id] = invoice_details  # invoice_details属性を更新

    def handle_api_error(self, error):
        QMessageBox.critical(self, "エラー", str(error))

    def parse_xml(self, xml_data):
        root = ET.fromstring(xml_data)

        namespaces = {
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        }

        invoice_id = root.find(".//cbc:ID", namespaces=namespaces).text
        sender_id = root.find(
            ".//cac:AccountingSupplierParty/cac:Party/cbc:EndpointID",
            namespaces=namespaces,
        ).text
        receiver_id = root.find(
            ".//cac:AccountingCustomerParty/cac:Party/cbc:EndpointID",
            namespaces=namespaces,
        ).text
        delivered_time = root.find(".//cbc:IssueDate", namespaces=namespaces).text

        invoice_line = root.find(".//cac:InvoiceLine", namespaces=namespaces)
        item_name = invoice_line.find(
            ".//cac:Item/cbc:Name", namespaces=namespaces
        ).text
        quantity = invoice_line.find(
            ".//cbc:InvoicedQuantity", namespaces=namespaces
        ).text
        price = invoice_line.find(
            ".//cac:Price/cbc:PriceAmount", namespaces=namespaces
        ).text

        invoice_details = {
            "invoice_id": invoice_id,
            "sender_id": sender_id,
            "receiver_id": receiver_id,
            "delivered_time": delivered_time,
            "item_name": item_name,
            "quantity": quantity,
            "price": price,
        }

        return invoice_details

    def parse_base64_xml(self, xml_data):
        decoded_xml_data = base64.b64decode(xml_data).decode("utf-8")
        root = ET.fromstring(decoded_xml_data)

        namespaces = {
            "cac": "urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2",
            "cbc": "urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2",
        }

        invoice_id = root.find(".//cbc:ID", namespaces=namespaces).text
        sender_id = root.find(
            ".//cac:AccountingSupplierParty/cac:Party/cbc:EndpointID",
            namespaces=namespaces,
        ).text
        receiver_id = root.find(
            ".//cac:AccountingCustomerParty/cac:Party/cbc:EndpointID",
            namespaces=namespaces,
        ).text
        delivered_time = root.find(".//cbc:IssueDate", namespaces=namespaces).text

        invoice_lines = root.findall(".//cac:InvoiceLine", namespaces=namespaces)

        invoice_details = {
            "invoice_id": invoice_id,
            "sender_id": sender_id,
            "receiver_id": receiver_id,
            "delivered_time": delivered_time,
            "items": [],
        }

        for line in invoice_lines:
            item_name = line.find(".//cac:Item/cbc:Name", namespaces=namespaces).text
            quantity = line.find(".//cbc:InvoicedQuantity", namespaces=namespaces).text
            price = line.find(
                ".//cac:Price/cbc:PriceAmount", namespaces=namespaces
            ).text

            item = {
                "item_name": item_name,
                "quantity": quantity,
                "price": price,
            }
            invoice_details["items"].append(item)

        return invoice_details

    def handle_base64_xml_result(self, result, invoice_id):
        print("base64処理を開始します")
        xml_data = result["payload"]
        print("XML Data:")  # デバッグ用の出力
        print(xml_data)  # デバッグ用の出力
        invoice_details = self.parse_base64_xml(xml_data)
        print(f"Parsed Invoice Details: {invoice_details}")  # デバッグ用の出力
        return invoice_details


class APICallThread(QThread):
    api_result = pyqtSignal(list, object)
    api_success = pyqtSignal(dict)
    api_error = pyqtSignal(Exception)

    def __init__(self, url, token):
        super().__init__()
        self.url = url
        self.token = token

    def run(self):
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Accept": "application/json",
        }

        try:
            response = requests.get(self.url, headers=headers)
            response.raise_for_status()  # ステータスコードが200番台以外の場合は例外を発生させる
            invoices = response.json()  # レスポンスのJSONデータを取得
            # print(f'invoices: {invoices}')
            self.api_result.emit(invoices, None)
        except Exception as e:
            self.api_result.emit([], e)


class SelectedAPICallThread(QThread):
    api_success = pyqtSignal(dict)
    api_error = pyqtSignal(Exception)

    def __init__(self, url, token):
        super().__init__()
        self.url = url
        self.token = token

    def run(self):
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Accept": "application/json",
        }

        try:
            response = requests.get(self.url, headers=headers)
            response.raise_for_status()  # ステータスコードが200番台以外の場合は例外を発生させる
            invoice = response.json()  # レスポンスのJSONデータを取得
            print(f"invoice: {invoice}")

            # invoiceのJSONからpayloadを取り出してprint
            payload = invoice["payload"]
            print(f"Payload: {payload}")

            self.api_success.emit(invoice)
        except Exception as e:
            self.api_error.emit(e)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    comparison_app = ComparisonApp()
    comparison_app.show()
    sys.exit(app.exec_())
